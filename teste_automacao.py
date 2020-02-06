"""
Script relativo ao teste de automação do candidato André França
Este scriot realiza o preenchimento de um formulário do Google.
Os dados estão contidos em uma planilha de excel.


Dependências:
$ pip install openpyxl
$ pip install requests
"""

from tkinter import *
import time
from tkinter.filedialog import askopenfilename
import os
from openpyxl import load_workbook
import requests

window = Tk()
window.title('Seja bem vindo!')
window.geometry('400x200')

lframe = Frame(window, borderwidth=2, relief="ridge", bg="white") # Caixa com contorno e cor
lframe.place(x=20, y=20, anchor="nw", width=360, height=40)
label2 = Label(lframe, text="", font=("Arial", 10), bg="white", justify='center') # label que mostra o nome do arquivo
label2.grid(column=0, row=0)

label2.configure(text='')


def click1():
    global file
    file = askopenfilename(initialdir=os.getcwd(), title="Selecione o arquivo.",
                           filetype=(("xlsx files", "*.xlsx"), ("all files", "*.*"))) # Seleção do arquivo
    label2.configure(text=str(os.path.basename(file))) # retorno do nome do aquivo
    return file


labelframe = Frame(window, borderwidth=1, relief="solid")
labelframe.place(x=140, y=80, anchor="nw", width=125, height=85)
btn1 = Button(labelframe, text='Escolher arquivo', command=click1)
btn1.grid(column=0, row=1, padx=10, pady=10)


def click2():
    global file
    wb = load_workbook(file)  # abrindo o Workbook test.xlsx
    ws = wb['Planilha1']  # selecionando a planilha 'Planilha1' dentro do Workbook test.xlsx

    linha = 0
    for line in ws:  # iterando em todas as linhas da 'Plan1'
        linha = linha + 1

    # Escrita no formulário online
    ### OBSERVAÇÃO ~> retire '/viewform' e add '/formResponse' ao fim do link ###

    url = "https://docs.google.com/forms/d/e/1FAIpQLSd3dOHHwAQ7JtxBc-Yyfu2B9jBT834Z6OLv_STMPm9qqYdyjg/formResponse"

    # Laço for com número de iterações = número de linhas do documento
    for x in range(1, linha + 1):
        nome = 'André José de França'
        Item = ws['A' + str(x)].value
        CEP = ws['C' + str(x)].value
        if CEP == None:
            CEP = 'NA'
        Endereco = ws['D' + str(x)].value
        # verifica se o Complemento está vazio
        if Endereco == None:
            Endereco = 'NA'
        Complemento = ws['E' + str(x)].value
        # verifica se o Complemento está vazio
        if Complemento == None:
            Complemento = 'NA'

        print(x) # verificar o funcionamento

        form_data = {'entry.366340186': nome, 'entry.1613519622': Item, 'entry.2045809580': CEP,
                     'entry.1334556551': Endereco, 'entry.1114882091': Complemento, 'draftResponse': [],
                     'pageHistory': 0}

        user_agent = {'Referer': url,
                      'User-Agent': "Mozilla/5.0 (X11; Linux i686) AppleWebKit/537.36 (KHTML, like Gecko) "
                                    "Chrome/28.0.1500.52 Safari/537.36"}
        r = requests.post(url, data=form_data, headers=user_agent)

        time.sleep(0.1) # para não sobrecarregar o site que contém o formulárioa

        if x == linha:
            label2.configure(text='Processo finalizado!') # informação de término do processo

# Posição do botão executar
btn2 = Button(labelframe, text='Executar', command=click2)
btn2.grid(column=0, row=2)

window.mainloop()
